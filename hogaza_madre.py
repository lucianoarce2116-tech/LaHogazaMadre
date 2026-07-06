"""
La Hogaza Madre — Sistema de Gestión Unificado
Versión 2.1 — Arquitectura modular con configuración externa JSON
Valores fijos por producto: redondeado, reinversión, ganancia
"""

import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
import re
import openpyxl
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES GLOBALES
# ══════════════════════════════════════════════════════════════════════════════
CONFIG_FILE       = "config_negocio.json"
CONFIG_DEFAULTS   = "config_default.json"
EXCEL_NAME        = "stock_la_hogaza_madre.xlsx"
VENTAS_HIST_FILE  = "ventas_historicas.json"
APP_VERSION       = "2.1"
REAJUSTE_MULTIPLO = 250

COLORS = {
    "bg_dark":    "#1A1412",
    "bg_mid":     "#241D1B",
    "bg_panel":   "#332B29",
    "bg_entry":   "#483E3C",
    "accent":     "#FF6F00",
    "green":      "#388E3C",
    "red":        "#C62828",
    "purple":     "#7B1FA2",
    "blue":       "#1976D2",
    "teal":       "#00897B",
    "fg_light":   "#FFF8F0",
    "fg_dim":     "#A6948E",
    "espera_bg":  "#0D47A1", "espera_fg":  "#BBDEFB",
    "armado_bg":  "#1B5E20", "armado_fg":  "#C8E6C9",
    "horno_bg":   "#B71C1C", "horno_fg":   "#FFCDD2",
    "listo_bg":   "#212121", "listo_fg":   "#9E9E9E",
}


# ══════════════════════════════════════════════════════════════════════════════
#  CAPA DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
class ConfigManager:
    """Maneja lectura y escritura del archivo JSON de configuración del negocio."""

    def __init__(self, defaults_file=CONFIG_DEFAULTS):
        self.data = {}
        self.defaults_file = defaults_file
        self.load()

    def _read_defaults(self):
        path = self.defaults_file
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                messagebox.showerror("Error crítico",
                    f"No se pudo leer '{path}'.\n"
                    f"Asegurate de que el archivo exista junto al programa.\n\nDetalle: {e}")
                return {}
        else:
            messagebox.showerror("Error crítico",
                f"Falta el archivo '{path}'.\n"
                "Reinstalá la aplicación o copiá el archivo de configuración por defecto.")
            return {}

    def load(self):
        defaults = self._read_defaults()
        if not defaults:
            self.data = {}
            return

        if not os.path.exists(CONFIG_FILE):
            self.data = defaults
            self.save()
        else:
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                for key, val in defaults.items():
                    if key not in self.data:
                        self.data[key] = val
            except Exception as e:
                messagebox.showerror("Error de configuración",
                    f"No se pudo leer {CONFIG_FILE}.\nSe usarán valores por defecto.\n\nDetalle: {e}")
                self.data = defaults

    def save(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo guardar {CONFIG_FILE}:\n{e}")
            return False

    @property
    def costos(self):    return self.data["costos_unidad"]
    @property
    def precios(self):   return self.data["precios_venta"]
    @property
    def recetas(self):   return self.data["recetas"]
    @property
    def grupos(self):    return self.data["grupos_menu"]
    @property
    def valores_fijos(self):
        if "valores_fijos" not in self.data:
            self.data["valores_fijos"] = {}
        return self.data["valores_fijos"]

    def get_valores_fijos(self, producto):
        return self.valores_fijos.get(producto, {"costo_redondeado": 0, "reinversion_fija": 0, "ganancia_fija": 0})

    def costo_receta(self, producto):
        receta = self.recetas.get(producto, {})
        return sum(cant * self.costos.get(ing, 0) for ing, cant in receta.items())

    def calcular_precio_venta(self, producto):
        vf = self.get_valores_fijos(producto)
        return vf["costo_redondeado"] + vf["reinversion_fija"] + vf["ganancia_fija"]

    def reajustar_redondeado(self, producto, multiplo=REAJUSTE_MULTIPLO):
        costo_real = self.costo_receta(producto)
        vf = self.get_valores_fijos(producto)
        if costo_real > vf["costo_redondeado"]:
            nuevo = ((costo_real // multiplo) + 1) * multiplo
            self.valores_fijos[producto]["costo_redondeado"] = nuevo
            self.precios[producto] = nuevo + vf["reinversion_fija"] + vf["ganancia_fija"]
            self.save()
            return True, nuevo, costo_real
        else:
            return False, vf["costo_redondeado"], costo_real


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS GLOBALES
# ══════════════════════════════════════════════════════════════════════════════
def validar_hora(s):
    return bool(re.fullmatch(r"\d{1,2}:\d{2}", s.strip()))

def hora_key(s):
    try:
        h, m = s.strip().split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 9999

def cargar_ventas_historicas():
    if not os.path.exists(VENTAS_HIST_FILE):
        return {"total": 0, "productos": {}}
    try:
        with open(VENTAS_HIST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"total": 0, "productos": {}}

def guardar_ventas_historicas(data):
    try:
        with open(VENTAS_HIST_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
#  PESTAÑA 1 — GESTIÓN DE PEDIDOS
# ══════════════════════════════════════════════════════════════════════════════
class PedidosTab(tk.Frame):

    def __init__(self, parent, cfg: ConfigManager, status_cb):
        super().__init__(parent, bg=COLORS["bg_mid"])
        self.cfg       = cfg
        self.status_cb = status_cb

        self.pedidos_totales           = []
        self.productos_pedido_actual   = []
        self.gastos_acumulados         = {k: 0.0 for k in cfg.costos}

        self._build_ui()

    def _build_ui(self):
        self._build_header()
        self._build_form()
        self._build_controles()
        self._build_tabla()
        self._build_footer()

    def _build_header(self):
        hdr = tk.Frame(self, bg=COLORS["bg_mid"], pady=6)
        hdr.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(hdr, text="Pedidos", font=("Segoe UI", 14, "bold"),
                 fg=COLORS["fg_light"], bg=COLORS["bg_mid"]).pack(side="left")
        self.lbl_turno = tk.Label(hdr, text="$0  ·  0 pedidos",
                                  font=("Segoe UI", 10), fg=COLORS["fg_dim"],
                                  bg=COLORS["bg_mid"])
        self.lbl_turno.pack(side="right")

    def _build_form(self):
        frm = tk.Frame(self, bg=COLORS["bg_panel"], padx=14, pady=12)
        frm.pack(fill="x", padx=16, pady=(8, 0))

        # Title inside card
        tk.Label(frm, text="Nuevo Pedido", font=("Segoe UI", 10, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_panel"]).pack(anchor="w", pady=(0, 8))

        row1 = tk.Frame(frm, bg=COLORS["bg_panel"])
        row1.pack(fill="x")
        # Col 1: cliente, hora, pago
        c1 = tk.Frame(row1, bg=COLORS["bg_panel"])
        c1.pack(side="left", fill="y", padx=(0, 16))
        for lbl in ("Cliente", "Hora (HH:MM)", "Pago"):
            tk.Label(c1, text=lbl, font=("Segoe UI", 9, "bold"),
                     bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w", pady=(4, 0))

        c1e = tk.Frame(row1, bg=COLORS["bg_panel"])
        c1e.pack(side="left", fill="y", padx=(8, 0))
        self.entry_nombre = tk.Entry(c1e, font=("Segoe UI", 10), width=18,
                                     bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                     insertbackground="white")
        self.entry_nombre.pack(pady=(2, 0))
        self.entry_hora = tk.Entry(c1e, font=("Segoe UI", 10), width=10,
                                    bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                    insertbackground="white")
        self.entry_hora.insert(0, "21:00")
        self.entry_hora.pack(pady=(4, 0), anchor="w")
        self.combo_pago = ttk.Combobox(c1e, values=["NO PAGO", "SÍ - PAGO"],
                                       state="readonly", width=12)
        self.combo_pago.current(0)
        self.combo_pago.pack(pady=(4, 0), anchor="w")

        # Col 2: producto, cantidad
        c2 = tk.Frame(row1, bg=COLORS["bg_panel"])
        c2.pack(side="left", fill="y", padx=(0, 16))
        tk.Label(c2, text="Variedad", font=("Segoe UI", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w", pady=(4, 0))
        self.combo_producto = ttk.Combobox(c2, state="readonly", width=30,
                                           style="Menu.TCombobox")
        self.combo_producto.pack(pady=(2, 0), fill="x")

        qty_frm = tk.Frame(c2, bg=COLORS["bg_panel"])
        qty_frm.pack(pady=(6, 0), anchor="w")
        tk.Label(qty_frm, text="Cant.", font=("Segoe UI", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(side="left")
        self.entry_cant = tk.Entry(qty_frm, font=("Segoe UI", 11), width=5, justify="center",
                                    bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                    insertbackground="white")
        self.entry_cant.insert(0, "1")
        self.entry_cant.pack(side="left", padx=(6, 4))
        tk.Button(qty_frm, text="+ Agregar", command=self.agregar_producto,
                  bg=COLORS["accent"], fg="white",
                  font=("Segoe UI", 9, "bold"), padx=10, pady=2).pack(side="left")
        tk.Button(qty_frm, text="↩ Quitar", command=self.quitar_ultimo,
                  bg=COLORS["bg_entry"], fg="white",
                  font=("Segoe UI", 8), padx=8).pack(side="left", padx=(4, 0))

        # Col 3: resumen + wsp
        c3 = tk.Frame(row1, bg=COLORS["bg_panel"])
        c3.pack(side="left", fill="both", expand=True)
        rsum = tk.Frame(c3, bg=COLORS["bg_panel"])
        rsum.pack(fill="both", expand=True)
        tk.Label(rsum, text="Resumen", font=("Segoe UI", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w", pady=(4, 0))
        self.txt_resumen = tk.Text(rsum, height=5, font=("Segoe UI", 9),
                                    bg=COLORS["bg_dark"], fg=COLORS["fg_light"],
                                    state="disabled", relief="flat", padx=6, pady=4)
        self.txt_resumen.pack(side="left", fill="both", expand=True, pady=(2, 0))
        tk.Button(rsum, text="📋\nCopiar\nWsp", command=self.copiar_wsp,
                  bg=COLORS["bg_entry"], fg="white",
                  font=("Segoe UI", 8, "bold"), width=4).pack(side="left", padx=(6, 0), pady=(2, 0))

        # Action buttons row
        acc = tk.Frame(frm, bg=COLORS["bg_panel"])
        acc.pack(fill="x", pady=(10, 0))
        tk.Button(acc, text="💾 GUARDAR PEDIDO",
                  command=self.guardar_pedido,
                  bg=COLORS["green"], fg="white",
                  font=("Segoe UI", 10, "bold"), pady=6).pack(side="left", fill="x", expand=True, padx=(0, 6))
        tk.Button(acc, text="Limpiar", command=self.limpiar_form,
                  bg=COLORS["bg_entry"], fg="white",
                  font=("Segoe UI", 9)).pack(side="left")

        self.refresh_combo_productos()

    def _build_controles(self):
        frm = tk.Frame(self, bg=COLORS["bg_mid"], pady=4)
        frm.pack(fill="x", padx=16)

        def btn(parent, text, cmd, bg):
            return tk.Button(parent, text=text, command=cmd, bg=bg,
                             fg="white", font=("Segoe UI", 8, "bold"), padx=6, pady=3)

        tk.Label(frm, text="Cocina:", font=("Segoe UI", 8, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_dim"]).pack(side="left")
        btn(frm, "Espera",   lambda: self.set_estado("espera"),    "#1565C0").pack(side="left", padx=2)
        btn(frm, "Armado",   lambda: self.set_estado("armado"),    "#2E7D32").pack(side="left", padx=2)
        btn(frm, "Horno",    lambda: self.set_estado("horno"),     "#C62828").pack(side="left", padx=2)
        btn(frm, "Entregado", lambda: self.set_estado("entregado"), "#424242").pack(side="left", padx=2)

        tk.Label(frm, text="  Caja:", font=("Segoe UI", 8, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_dim"]).pack(side="left", padx=(12, 0))
        btn(frm, "Pagado", lambda: self.set_pago("SÍ - PAGO"), COLORS["teal"]).pack(side="left", padx=2)
        btn(frm, "No Pago", lambda: self.set_pago("NO PAGO"), "#AD1457").pack(side="left", padx=2)

        tk.Label(frm, text="  Envío:", font=("Segoe UI", 8, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_dim"]).pack(side="left", padx=(12, 0))
        btn(frm, "Delivery",       lambda: self.set_envio("Delivery"),         "#E65100").pack(side="left", padx=2)
        btn(frm, "Retira",         lambda: self.set_envio("Retira en local"), "#0277BD").pack(side="left", padx=2)

        tk.Label(frm, text="  Pedido:", font=("Segoe UI", 8, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_dim"]).pack(side="left", padx=(12, 0))
        btn(frm, "✕ Cancelar", self.cancelar_pedido, COLORS["red"]).pack(side="left", padx=2)

    def _build_tabla(self):
        frm = tk.Frame(self, bg=COLORS["bg_panel"], padx=2, pady=2)
        frm.pack(fill="both", expand=True, padx=16, pady=(6, 4))

        hdr_tbl = tk.Frame(frm, bg=COLORS["bg_panel"])
        hdr_tbl.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(hdr_tbl, text="Órdenes del Turno", font=("Segoe UI", 10, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_panel"]).pack(side="left")

        cols = ("hora", "nombre", "detalle", "total", "pago", "envio")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings")
        for col, txt, w, anch in [
            ("hora",    "Hora",      88,  "center"),
            ("nombre",  "Cliente",   130,  "w"),
            ("detalle", "Detalle",   200, "w"),
            ("total",   "Total",      95,  "center"),
            ("pago",    "Caja",       95,  "center"),
            ("envio",   "Envío",     135,  "center"),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor=anch)
        self.tree.column("detalle", stretch=True)

        self.tree.tag_configure("espera",    background=COLORS["espera_bg"], foreground=COLORS["espera_fg"])
        self.tree.tag_configure("armado",    background=COLORS["armado_bg"], foreground=COLORS["armado_fg"])
        self.tree.tag_configure("horno",     background=COLORS["horno_bg"],  foreground=COLORS["horno_fg"])
        self.tree.tag_configure("entregado", background=COLORS["listo_bg"],  foreground=COLORS["listo_fg"])

        sb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        sb.pack(side="right", fill="y", pady=(0, 8))

    def _build_footer(self):
        tk.Button(self, text="Finalizar Servicio — Restar Stock & Generar Balance",
                  command=self.finalizar_servicio,
                  bg=COLORS["purple"], fg="white",
                  font=("Segoe UI", 10, "bold"), pady=8).pack(fill="x", padx=16, pady=(4, 10))

    def refresh_combo_productos(self):
        grupos = self.cfg.grupos
        labels = [g[0] for g in grupos]
        self.label_a_prod = {g[0]: g[1] for g in grupos}
        self.combo_producto["values"] = labels
        for i, g in enumerate(grupos):
            if g[1] is not None:
                self.combo_producto.current(i)
                break

    def _prod_actual(self):
        label = self.combo_producto.get()
        prod  = self.label_a_prod.get(label)
        if prod is None:
            messagebox.showwarning("Aviso", "Seleccioná un producto válido, no un separador.")
        return prod

    def agregar_producto(self):
        prod = self._prod_actual()
        if not prod:
            return
        try:
            cant = int(self.entry_cant.get().strip())
            if cant <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Cantidad inválida.")
            return
        precio   = self.cfg.precios.get(prod, 0)
        subtotal = precio * cant
        self.productos_pedido_actual.append({"prod": prod, "cant": cant, "subtotal": subtotal})
        self._refresh_resumen()
        self.entry_cant.delete(0, tk.END)
        self.entry_cant.insert(0, "1")
        self.combo_producto.focus_set()

    def quitar_ultimo(self):
        if not self.productos_pedido_actual:
            messagebox.showwarning("Aviso", "No hay productos para quitar.")
            return
        el = self.productos_pedido_actual.pop()
        self._refresh_resumen()
        self.status_cb(f"Eliminado: {el['cant']}x {el['prod']}", COLORS["red"])

    def _refresh_resumen(self):
        self.txt_resumen.config(state="normal")
        self.txt_resumen.delete("1.0", tk.END)
        total = 0
        for it in self.productos_pedido_actual:
            self.txt_resumen.insert(tk.END,
                f"  {it['cant']}x {it['prod']}  (${it['subtotal']:,.0f})\n")
            total += it["subtotal"]
        if self.productos_pedido_actual:
            self.txt_resumen.insert(tk.END, f"\nTOTAL: ${total:,.0f}")
        self.txt_resumen.config(state="disabled")
        self._refresh_header()

    def _refresh_header(self):
        total = sum(p["total"] for p in self.pedidos_totales)
        cant  = len(self.pedidos_totales)
        self.lbl_turno.config(text=f"${total:,.0f}  |  {cant} pedidos")

    def copiar_wsp(self):
        if not self.productos_pedido_actual:
            messagebox.showwarning("Aviso", "No hay productos en el pedido.")
            return
        nombre = self.entry_nombre.get().strip() or "Particular"
        hora   = self.entry_hora.get().strip()
        total  = sum(i["subtotal"] for i in self.productos_pedido_actual)
        txt    = (f"*LA HOGAZA MADRE - CONFIRMACIÓN*\n\n"
                  f"Cliente: {nombre}\nHora: {hora}\n"
                  f"{'─'*40}\n")
        for it in self.productos_pedido_actual:
            txt += f"{it['cant']}x {it['prod']}\n"
        txt += f"{'─'*40}\nTotal: ${total:,.0f}\n\nGracias por elegirnos!"
        self.winfo_toplevel().clipboard_clear()
        self.winfo_toplevel().clipboard_append(txt)
        self.status_cb("Texto copiado para WhatsApp", COLORS["teal"])

    def guardar_pedido(self):
        nombre = self.entry_nombre.get().strip()
        hora   = self.entry_hora.get().strip()
        pago   = self.combo_pago.get()
        if not nombre:
            messagebox.showwarning("Aviso", "Falta el nombre del cliente.")
            self.entry_nombre.focus_set(); return
        if not validar_hora(hora):
            messagebox.showwarning("Aviso", "Hora inválida — usá HH:MM (ej: 20:30).")
            self.entry_hora.focus_set(); return
        if not self.productos_pedido_actual:
            messagebox.showwarning("Aviso", "No hay productos en el pedido."); return

        total = sum(i["subtotal"] for i in self.productos_pedido_actual)
        det   = ", ".join(f"{i['cant']}x {i['prod']}" for i in self.productos_pedido_actual)

        for it in self.productos_pedido_actual:
            for ing, q in self.cfg.recetas.get(it["prod"], {}).items():
                self.gastos_acumulados[ing] = self.gastos_acumulados.get(ing, 0) + q * it["cant"]

        self.pedidos_totales.append({
            "nombre": nombre, "hora": hora, "pago": pago,
            "envio": "Por definir", "detalle": det,
            "total": total, "estado": "espera",
            "items": list(self.productos_pedido_actual)
        })
        self.pedidos_totales.sort(key=lambda x: hora_key(x["hora"]))
        self._refresh_tabla()
        self._refresh_header()
        self.limpiar_form()
        self.status_cb(f"Pedido de {nombre} guardado — ${total:,.0f}", COLORS["green"])

    def limpiar_form(self):
        self.productos_pedido_actual = []
        self.entry_nombre.delete(0, tk.END)
        self.entry_hora.delete(0, tk.END)
        self.entry_hora.insert(0, "21:00")
        self.combo_pago.current(0)
        self.refresh_combo_productos()
        self.entry_cant.delete(0, tk.END)
        self.entry_cant.insert(0, "1")
        self.txt_resumen.config(state="normal")
        self.txt_resumen.delete("1.0", tk.END)
        self.txt_resumen.config(state="disabled")
        self.entry_nombre.focus_set()

    def _refresh_tabla(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in self.pedidos_totales:
            self.tree.insert("", "end", values=(
                p["hora"], p["nombre"], p["detalle"],
                f"${p['total']:,.0f}", p["pago"], p["envio"]
            ), tags=(p["estado"],))

    def _ped_seleccionado(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccioná un pedido de la lista primero.")
            return None
        v = self.tree.item(sel)["values"]
        for p in self.pedidos_totales:
            if p["nombre"] == v[1] and p["hora"] == str(v[0]):
                return p
        return None

    def set_estado(self, estado):
        p = self._ped_seleccionado()
        if p:
            p["estado"] = estado
            self._refresh_tabla()

    def set_pago(self, pago):
        p = self._ped_seleccionado()
        if p:
            p["pago"] = pago
            self._refresh_tabla()

    def set_envio(self, envio):
        p = self._ped_seleccionado()
        if p:
            p["envio"] = envio
            self._refresh_tabla()

    def cancelar_pedido(self):
        p = self._ped_seleccionado()
        if not p:
            return
        if not messagebox.askyesno("Cancelar pedido",
            f"Cancelar el pedido de {p['nombre']} ({p['hora']})?\n"
            "Se revertirá el uso de insumos."):
            return
        for it in p.get("items", []):
            for ing, q in self.cfg.recetas.get(it["prod"], {}).items():
                self.gastos_acumulados[ing] = max(
                    0, self.gastos_acumulados.get(ing, 0) - q * it["cant"])
        self.pedidos_totales.remove(p)
        self._refresh_tabla()
        self._refresh_header()
        self.status_cb(f"Pedido de {p['nombre']} cancelado", COLORS["red"])

    def finalizar_servicio(self):
        if not self.pedidos_totales:
            messagebox.showwarning("Aviso", "No hay pedidos en este turno."); return
        if not messagebox.askyesno("Finalizar servicio",
            "Cerras el servicio?\n\n"
            "Se descontará el stock del Excel\n"
            "Se generará el balance del turno\n\n"
            "Esta acción no se puede deshacer."):
            return
        if not os.path.exists(EXCEL_NAME):
            messagebox.showerror("Error",
                f"No se encuentra '{EXCEL_NAME}'.\n"
                "Asegurate de que el archivo de stock esté junto al programa.")
            return

        try:
            wb = openpyxl.load_workbook(EXCEL_NAME)
            ws = wb["Stock General"]
            fila = 4
            while True:
                prod = ws.cell(row=fila, column=1).value
                if prod is None:
                    break
                ini  = ws.cell(row=fila, column=2).value or 0
                ocu  = self.gastos_acumulados.get(str(prod), 0.0)
                cu   = self.cfg.costos.get(str(prod), 0.0)
                ws.cell(row=fila, column=3, value=ocu)
                ws.cell(row=fila, column=4, value=ocu * cu)
                ws.cell(row=fila, column=5, value=max(0, ini - ocu))
                fila += 1
            wb.save(EXCEL_NAME)

            ahora = datetime.now()
            L = "=" * 62
            S = "-" * 62
            r  = ""

            # ── BLOQUE 1: DETALLE DE PEDIDOS ──
            r += "  DETALLE DE PEDIDOS\n" + S + "\n"
            estados_txt = {"espera":"En espera","armado":"Armado","horno":"En horno","entregado":"Entregado"}
            for idx, p in enumerate(self.pedidos_totales, 1):
                r += (f"\n  {idx:>2}. {p['nombre']:<20}  {p['hora']}  |  "
                      f"{p['pago']}  |  {p['envio']}  |  {estados_txt.get(p['estado'], p['estado'])}\n")
                for it in p.get("items", []):
                    pv = self.cfg.precios.get(it["prod"], 0)
                    r += f"    • {it['cant']}u × {it['prod']} → ${pv * it['cant']:,.0f}\n"
                r += f"    Subtotal pedido: ${p['total']:,.0f}\n"
            r += "\n" + L + "\n\n"

            # ── BLOQUE 2: TOTAL DE COSTOS DEL SERVICIO ──
            r += "  TOTAL DE COSTOS DEL SERVICIO\n" + S + "\n"
            r += f"  {'PRODUCTO':<38}{'CANT':>5}  {'COSTO UNIT':>10}  {'COSTO TOTAL':>12}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*10}  {'─'*12}\n"
            costo_x_prod = {prod: self.cfg.costo_receta(prod) for prod in self.cfg.precios}
            total_costo_real = 0
            for prod in sorted(costo_x_prod.keys()):
                cant = sum(it["cant"] for p in self.pedidos_totales for it in p.get("items", []) if it["prod"] == prod)
                if cant == 0:
                    continue
                costo_u = costo_x_prod[prod]
                costo_t = costo_u * cant
                total_costo_real += costo_t
                vf = self.cfg.get_valores_fijos(prod)
                redondeado = vf["costo_redondeado"]
                r += f"  {prod:<38}{cant:>5}  ${redondeado:>8,.0f}  ${costo_t:>10,.0f}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*10}  {'─'*12}\n"
            r += f"  {'COSTO TOTAL DEL SERVICIO':<54}  ${total_costo_real:>10,.0f}\n\n"
            r += L + "\n\n"

            # ── BLOQUE 3: TOTAL DE VENTAS POR PRODUCTO ──
            r += "  TOTAL DE VENTAS POR PRODUCTO\n" + S + "\n"
            r += f"  {'PRODUCTO':<38}{'CANT':>5}  {'P.VENTA':>10}  {'TOTAL VENDIDO':>14}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*10}  {'─'*14}\n"
            total_ventas = 0
            for prod in sorted(self.cfg.precios.keys()):
                cant = sum(it["cant"] for p in self.pedidos_totales for it in p.get("items", []) if it["prod"] == prod)
                if cant == 0:
                    continue
                pv = self.cfg.precios.get(prod, 0)
                subtotal = pv * cant
                total_ventas += subtotal
                r += f"  {prod:<38}{cant:>5}  ${pv:>8,.0f}  ${subtotal:>12,.0f}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*10}  {'─'*14}\n"
            r += f"  {'TOTAL VENTAS DEL SERVICIO':<54}  ${total_ventas:>12,.0f}\n\n"
            r += L + "\n\n"

            # ── BLOQUE 4: TOTAL DE GANANCIAS ──
            r += "  TOTAL DE GANANCIAS\n" + S + "\n"
            r += f"  {'PRODUCTO':<38}{'CANT':>5}  {'GANANCIA U':>11}  {'GANANCIA TOTAL':>15}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*11}  {'─'*15}\n"
            total_ganancia = 0
            for prod in sorted(self.cfg.precios.keys()):
                cant = sum(it["cant"] for p in self.pedidos_totales for it in p.get("items", []) if it["prod"] == prod)
                if cant == 0:
                    continue
                vf = self.cfg.get_valores_fijos(prod)
                gan_u = vf["ganancia_fija"]
                gan_t = gan_u * cant
                total_ganancia += gan_t
                r += f"  {prod:<38}{cant:>5}  ${gan_u:>9,.0f}  ${gan_t:>13,.0f}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*11}  {'─'*15}\n"
            r += f"  {'GANANCIA TOTAL DEL SERVICIO':<54}  ${total_ganancia:>13,.0f}\n\n"
            r += L + "\n\n"

            # ── BLOQUE 5: REINVERSIÓN ──
            r += "  REINVERSIÓN\n" + S + "\n"
            r += f"  {'PRODUCTO':<38}{'CANT':>5}  {'REINV. U':>9}  {'REINV. TOTAL':>13}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*9}  {'─'*13}\n"
            total_reinversion = 0
            for prod in sorted(self.cfg.precios.keys()):
                cant = sum(it["cant"] for p in self.pedidos_totales for it in p.get("items", []) if it["prod"] == prod)
                if cant == 0:
                    continue
                vf = self.cfg.get_valores_fijos(prod)
                reinv_u = vf["reinversion_fija"]
                reinv_t = reinv_u * cant
                total_reinversion += reinv_t
                r += f"  {prod:<38}{cant:>5}  ${reinv_u:>7,.0f}  ${reinv_t:>11,.0f}\n"
            r += f"  {'─'*38}{'─'*5}  {'─'*9}  {'─'*13}\n"
            r += f"  {'REINVERSIÓN TOTAL DEL SERVICIO':<54}  ${total_reinversion:>11,.0f}\n\n"
            r += L + "\n\n"

            # ── RESUMEN FINAL ──
            r += "  RESUMEN FINAL DEL TURNO\n" + S + "\n"
            r += f"  Total ventas:                        ${total_ventas:>12,.0f}\n"
            r += f"  Total costos reales (receta):        ${total_costo_real:>12,.0f}\n"
            r += f"  Total reinversión:                   ${total_reinversion:>12,.0f}\n"
            r += f"  Total ganancia socios:               ${total_ganancia:>12,.0f}\n"
            r += L + "\n"
            r += f"  Sistema La Hogaza Madre v{APP_VERSION} — {ahora.strftime('%d/%m/%Y %H:%M')}\n"
            r += L + "\n"

            os.makedirs("balances", exist_ok=True)
            nombre_arch = os.path.join("balances", f"Balance_{ahora.strftime('%d-%m-%Y')}.txt")
            try:
                with open(nombre_arch, "w", encoding="utf-8") as f:
                    f.write(r)
                path_av = nombre_arch
            except Exception:
                path_av = os.path.join(os.path.expanduser("~"), "Desktop", f"Balance_{ahora.strftime('%d-%m-%Y')}.txt")
                with open(path_av, "w", encoding="utf-8") as f:
                    f.write(r)

            # ── GUARDAR VENTAS HISTORICAS ──
            hist = cargar_ventas_historicas()
            hist["total"] += total_ventas
            for p in self.pedidos_totales:
                for it in p.get("items", []):
                    prod = it["prod"]
                    hist["productos"][prod] = hist["productos"].get(prod, 0) + it["cant"]
            guardar_ventas_historicas(hist)

            messagebox.showinfo("Servicio cerrado",
                f"Balance generado y stock actualizado.\n\n"
                f"Total ventas: ${total_ventas:,.0f}\n"
                f"Total ganancia socios: ${total_ganancia:,.0f}\n\n"
                f"Archivo: {path_av}")

        except Exception as e:
            messagebox.showerror("Error", f"Problema al cerrar el servicio:\n{e}")


# ══════════════════════════════════════════════════════════════════════════════
#  PESTAÑA 2 — CONTROL DE STOCK
# ══════════════════════════════════════════════════════════════════════════════
class StockTab(tk.Frame):

    def __init__(self, parent, cfg: ConfigManager, status_cb):
        super().__init__(parent, bg=COLORS["bg_mid"])
        self.cfg       = cfg
        self.status_cb = status_cb
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=COLORS["bg_mid"], pady=6)
        hdr.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(hdr, text="Stock", font=("Segoe UI", 14, "bold"),
                 fg=COLORS["fg_light"], bg=COLORS["bg_mid"]).pack(side="left")
        tk.Button(hdr, text="🔄 Actualizar", command=self.cargar_stock,
                  bg=COLORS["blue"], fg="white",
                  font=("Segoe UI", 9, "bold"), padx=10).pack(side="right")

        frm = tk.Frame(self, bg=COLORS["bg_panel"], padx=2, pady=2)
        frm.pack(fill="both", expand=True, padx=16, pady=(8, 4))

        hdr_tbl = tk.Frame(frm, bg=COLORS["bg_panel"])
        hdr_tbl.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(hdr_tbl, text="Stock Actual (desde Excel)", font=("Segoe UI", 10, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_panel"]).pack(side="left")

        cols = ("producto", "inicio", "ocupado", "costo_ocu", "restante")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings")
        for col, txt, w, anch in [
            ("producto",  "Producto",            200, "w"),
            ("inicio",    "Cant. Inicio",         120, "center"),
            ("ocupado",   "Cant. Ocupada",        120, "center"),
            ("costo_ocu", "Costo Ocupado ($)",    150, "center"),
            ("restante",  "Restante",             120, "center"),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor=anch)

        self.tree.tag_configure("bajo",  background="#4A1010", foreground="#FFCDD2")
        self.tree.tag_configure("medio", background="#4A3800", foreground="#FFF9C4")
        self.tree.tag_configure("ok",    background="#1B3A1F", foreground="#C8E6C9")

        sb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        sb.pack(side="right", fill="y", pady=(0, 8))

        self.frm_entradas = tk.Frame(self, bg=COLORS["bg_panel"], padx=14, pady=10)
        self.frm_entradas.pack(fill="x", padx=16, pady=(0, 10))
        tk.Label(self.frm_entradas, text="Cargar Stock Inicial",
                 font=("Segoe UI", 10, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_panel"]).pack(anchor="w", pady=(0, 6))
        self._reconstruir_entradas()

        self.cargar_stock()

    def _reconstruir_entradas(self, items=None):
        """Crea/actualiza la grilla de entradas para carga de stock inicial (con scroll)."""
        for w in self.frm_entradas.winfo_children():
            w.destroy()

        # Canvas + Scrollbar para que el contenido sea desplazable
        canvas = tk.Canvas(self.frm_entradas, borderwidth=0, highlightthickness=0,
                           bg=COLORS["bg_panel"])
        sb = ttk.Scrollbar(self.frm_entradas, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=COLORS["bg_panel"])
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(1, width=e.width))
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        # Scroll con rueda del mouse (solo sobre el canvas)
        def _on_mw(ev):
            canvas.yview_scroll(int(-1 * ev.delta / 120), "units")
        canvas.bind("<MouseWheel>", _on_mw)
        inner.bind("<MouseWheel>", _on_mw)

        if items is None:
            datos = self._leer_todo_excel()
            items = sorted(set(datos.keys()) | set(self.cfg.costos.keys()),
                           key=lambda x: list(self.cfg.costos.keys()).index(x) if x in self.cfg.costos else 999)

        # 4 grupos de columnas para ocupar mejor el ancho y reducir altura
        num_grupos = 4
        total_cols = num_grupos * 2

        tk.Label(inner, text="Suma los valores ingresados al restante actual del Excel.",
                 font=("Arial", 8), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_panel"]).grid(row=0, column=0, columnspan=total_cols, sticky="w", pady=(0, 6))

        self.entradas = {}
        filas_por_grupo = (len(items) + num_grupos - 1) // num_grupos
        for idx, ins in enumerate(items):
            col_group = idx % num_grupos
            col_base = col_group * 2
            row = idx // num_grupos + 1
            tk.Label(inner, text=f"{ins}:",
                     font=("Arial", 8, "bold"), width=18, anchor="w",
                     bg=COLORS["bg_panel"], fg=COLORS["fg_light"]).grid(
                     row=row, column=col_base, padx=(8, 2), pady=2, sticky="w")
            e = tk.Entry(inner, font=("Arial", 9), width=10, justify="center",
                         bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                         insertbackground="white")
            e.insert(0, "0")
            e.grid(row=row, column=col_base+1, padx=(0, 12), pady=2)
            self.entradas[ins] = e

        btn_row = filas_por_grupo + 1
        tk.Button(inner, text="GUARDAR STOCK INICIAL EN EXCEL",
                  command=self.guardar_stock,
                  bg=COLORS["green"], fg="white",
                  font=("Arial", 10, "bold"), pady=5).grid(
                  row=btn_row, column=0, columnspan=total_cols, sticky="ew", padx=8, pady=(10, 0))

    def _leer_todo_excel(self):
        """Lee todas las filas del Excel y devuelve {producto: {fila, inicio, ocupado, costo_ocu, restante}}."""
        datos = {}
        if not os.path.exists(EXCEL_NAME):
            return datos
        try:
            wb = openpyxl.load_workbook(EXCEL_NAME, data_only=True)
            ws = wb["Stock General"]
            fila = 4
            while True:
                prod = ws.cell(row=fila, column=1).value
                if prod is None:
                    break
                prod = str(prod)
                datos[prod] = {
                    "fila": fila,
                    "inicio":   float(ws.cell(row=fila, column=2).value or 0),
                    "ocupado":  float(ws.cell(row=fila, column=3).value or 0),
                    "costo_ocu": float(ws.cell(row=fila, column=4).value or 0),
                    "restante": float(ws.cell(row=fila, column=5).value or 0),
                }
                fila += 1
            wb.close()
        except Exception:
            pass
        return datos

    def cargar_stock(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        datos = self._leer_todo_excel()
        if not datos:
            self.status_cb(f"No se encuentra {EXCEL_NAME} o esta vacio", COLORS["red"])
            return
        for prod, d in datos.items():
            pct = (d["restante"] / d["inicio"] * 100) if d["inicio"] > 0 else 100
            tag = "bajo" if pct < 20 else ("medio" if pct < 50 else "ok")
            self.tree.insert("", "end", values=(
                prod,
                f"{d['inicio']:g}", f"{d['ocupado']:g}",
                f"${d['costo_ocu']:,.2f}", f"{d['restante']:g}"
            ), tags=(tag,))
        # Reconstruir entradas si hay productos nuevos en el Excel
        items_actuales = set(self.entradas.keys())
        items_excel = set(datos.keys())
        if items_excel - items_actuales:
            self._reconstruir_entradas(items=sorted(items_actuales | items_excel))
        self.status_cb("Stock actualizado desde Excel", COLORS["teal"])

    def guardar_stock(self):
        if not os.path.exists(EXCEL_NAME):
            self._crear_excel()
        try:
            datos = self._leer_todo_excel()
            wb = openpyxl.load_workbook(EXCEL_NAME)
            ws = wb["Stock General"]

            # Obtener todos los productos entre Excel y costos
            todos = set(datos.keys()) | set(self.cfg.costos.keys()) | set(self.entradas.keys())
            for prod in sorted(todos):
                fila = datos[prod]["fila"] if prod in datos else None
                if fila is None:
                    # Nuevo producto: agregar al final
                    fila = max((d["fila"] for d in datos.values()), default=3) + 1
                    ws.cell(row=fila, column=1, value=prod)
                try:
                    nuevo = float(self.entradas.get(prod, tk.Entry()).get().strip() or 0)
                except ValueError:
                    nuevo = 0.0
                restante_anterior = datos[prod]["restante"] if prod in datos else 0.0
                total = restante_anterior + nuevo
                ws.cell(row=fila, column=2, value=total)
                ws.cell(row=fila, column=3, value=0.0)
                ws.cell(row=fila, column=4, value=0.0)
                ws.cell(row=fila, column=5, value=total)
            wb.save(EXCEL_NAME)
            for e in self.entradas.values():
                e.delete(0, tk.END)
                e.insert(0, "0")
            self.cargar_stock()
            messagebox.showinfo("Éxito", "Stock guardado sumando restantes anteriores.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el stock:\n{e}")

    def _leer_restantes(self):
        datos = self._leer_todo_excel()
        return {p: d["restante"] for p, d in datos.items()}

    def _crear_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock General"
        ws.append([])
        ws.append(["stock"])
        ws.append(["producto", "cantidad de inicio", "cantidad ocupada",
                   "costo de cantidad ocupada", "cantidad restante"])
        for ins in self.cfg.costos:
            ws.append([ins, 0, 0, 0, 0])
        wb.save(EXCEL_NAME)


# ══════════════════════════════════════════════════════════════════════════════
#  PESTAÑA 3 — CONFIGURACIÓN DE CARTA Y RECETAS
# ══════════════════════════════════════════════════════════════════════════════
class ConfigTab(tk.Frame):

    def __init__(self, parent, cfg: ConfigManager, status_cb, refresh_cb):
        super().__init__(parent, bg=COLORS["bg_mid"])
        self.cfg        = cfg
        self.status_cb  = status_cb
        self.refresh_cb = refresh_cb
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=COLORS["bg_mid"], pady=6)
        hdr.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(hdr, text="Configuración", font=("Segoe UI", 14, "bold"),
                 fg=COLORS["fg_light"], bg=COLORS["bg_mid"]).pack(side="left")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=16, pady=(8, 4))

        self.tab_costos   = tk.Frame(nb, bg=COLORS["bg_mid"])
        self.tab_precios  = tk.Frame(nb, bg=COLORS["bg_mid"])
        self.tab_recetas  = tk.Frame(nb, bg=COLORS["bg_mid"])

        nb.add(self.tab_costos,  text="  Costos de Insumos  ")
        nb.add(self.tab_precios, text="  Precios de Venta  ")
        nb.add(self.tab_recetas, text="  Recetas / Gramajes  ")

        self._build_costos()
        self._build_precios()
        self._build_recetas()

        tk.Button(self, text="💾 Guardar Configuración y Aplicar",
                  command=self.guardar_config,
                  bg=COLORS["purple"], fg="white",
                  font=("Segoe UI", 10, "bold"), pady=8).pack(
                  fill="x", padx=16, pady=(0, 10))

    def _build_costos(self):
        tk.Label(self.tab_costos,
                 text="Seleccioná un insumo y modificá su costo por unidad/gramo.",
                 font=("Arial", 9), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_mid"]).pack(padx=10, pady=(8, 4), anchor="w")

        pane = tk.Frame(self.tab_costos, bg=COLORS["bg_mid"])
        pane.pack(fill="both", expand=True, padx=10, pady=4)

        cols = ("insumo", "costo_actual")
        self.tree_costos = ttk.Treeview(pane, columns=cols, show="headings", height=16)
        self.tree_costos.heading("insumo",      text="Insumo")
        self.tree_costos.heading("costo_actual", text="Costo actual ($/u.g)")
        self.tree_costos.column("insumo",       width=220, anchor="w")
        self.tree_costos.column("costo_actual", width=160, anchor="center")
        sb = ttk.Scrollbar(pane, orient="vertical", command=self.tree_costos.yview)
        self.tree_costos.configure(yscrollcommand=sb.set)
        self.tree_costos.pack(side="left", fill="both")
        sb.pack(side="left", fill="y")
        self.tree_costos.bind("<<TreeviewSelect>>", self._on_select_costo)

        edit = tk.Frame(pane, bg=COLORS["bg_panel"], padx=16, pady=16)
        edit.pack(side="left", fill="y", padx=(16, 0))

        tk.Label(edit, text="Insumo seleccionado:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w")
        self.lbl_costo_ins = tk.Label(edit, text="—",
                                       font=("Arial", 11, "bold"),
                                       bg=COLORS["bg_panel"], fg=COLORS["fg_light"])
        self.lbl_costo_ins.pack(anchor="w", pady=(0, 10))

        tk.Label(edit, text="Nuevo costo ($):",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_light"]).pack(anchor="w")
        self.entry_nuevo_costo = tk.Entry(edit, font=("Arial", 12), width=14,
                                          justify="center",
                                          bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                          insertbackground="white")
        self.entry_nuevo_costo.pack(pady=6)

        tk.Button(edit, text="Actualizar costo",
                  command=self._actualizar_costo,
                  bg=COLORS["teal"], fg="white",
                  font=("Arial", 10, "bold"), pady=6).pack(fill="x")

        self._reload_tree_costos()

    def _reload_tree_costos(self):
        for r in self.tree_costos.get_children():
            self.tree_costos.delete(r)
        for ins, val in self.cfg.costos.items():
            self.tree_costos.insert("", "end", values=(ins, f"${val:g}"))

    def _on_select_costo(self, _=None):
        sel = self.tree_costos.selection()
        if not sel:
            return
        ins = self.tree_costos.item(sel)["values"][0]
        self.lbl_costo_ins.config(text=ins)
        self.entry_nuevo_costo.delete(0, tk.END)
        self.entry_nuevo_costo.insert(0, str(self.cfg.costos.get(ins, 0)))

    def _actualizar_costo(self):
        sel = self.tree_costos.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccioná un insumo primero."); return
        ins = self.tree_costos.item(sel)["values"][0]
        try:
            nuevo = float(self.entry_nuevo_costo.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Ingresá un número válido."); return
        self.cfg.costos[ins] = nuevo
        self._reload_tree_costos()
        if self._prod_precio_actual:
            self._actualizar_desglose_precio()
        self.status_cb(f"Costo de '{ins}' actualizado a ${nuevo:g} (sin guardar aún)", COLORS["accent"])

    def _build_precios(self):
        tk.Label(self.tab_precios,
                 text="Seleccioná un producto para ver y editar sus valores fijos.",
                 font=("Arial", 9), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_mid"]).pack(padx=10, pady=(8, 4), anchor="w")

        pane = tk.Frame(self.tab_precios, bg=COLORS["bg_mid"])
        pane.pack(fill="both", expand=True, padx=10, pady=4)

        cols = ("producto", "precio")
        self.tree_precios = ttk.Treeview(pane, columns=cols, show="headings", height=16)
        self.tree_precios.heading("producto", text="Producto")
        self.tree_precios.heading("precio",   text="Precio venta ($)")
        self.tree_precios.column("producto",  width=280, anchor="w")
        self.tree_precios.column("precio",    width=140, anchor="center")
        sb = ttk.Scrollbar(pane, orient="vertical", command=self.tree_precios.yview)
        self.tree_precios.configure(yscrollcommand=sb.set)
        self.tree_precios.pack(side="left", fill="both")
        sb.pack(side="left", fill="y")
        self.tree_precios.bind("<<TreeviewSelect>>", self._on_select_precio)

        edit = tk.Frame(pane, bg=COLORS["bg_panel"], padx=16, pady=16)
        edit.pack(side="left", fill="y", padx=(16, 0))

        # ── DESGLOSE DEL PRODUCTO SELECCIONADO ──
        self._prod_precio_actual = None

        tk.Label(edit, text="Producto seleccionado:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w")
        self.lbl_precio_prod = tk.Label(edit, text="—",
                                         font=("Arial", 11, "bold"),
                                         bg=COLORS["bg_panel"], fg=COLORS["fg_light"],
                                         wraplength=200)
        self.lbl_precio_prod.pack(anchor="w", pady=(0, 8))

        sep1 = tk.Frame(edit, height=1, bg=COLORS["fg_dim"])
        sep1.pack(fill="x", pady=4)

        frm_desg = tk.Frame(edit, bg=COLORS["bg_panel"])
        frm_desg.pack(fill="x", pady=2)

        rw = 0
        tk.Label(frm_desg, text="Costo real (receta):",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).grid(row=rw, column=0, sticky="w", pady=2)
        self.lbl_costo_real = tk.Label(frm_desg, text="$0",
                                        font=("Arial", 10, "bold"),
                                        bg=COLORS["bg_panel"], fg=COLORS["fg_light"])
        self.lbl_costo_real.grid(row=rw, column=1, padx=8, sticky="w")
        rw += 1

        tk.Label(frm_desg, text="Costo redondeado:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).grid(row=rw, column=0, sticky="w", pady=2)
        self.entry_redondeado = tk.Entry(frm_desg, font=("Arial", 10), width=12, justify="center",
                                          bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                          insertbackground="white")
        self.entry_redondeado.grid(row=rw, column=1, padx=8, pady=2, sticky="w")
        rw += 1

        tk.Label(frm_desg, text="Reinversión fija:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).grid(row=rw, column=0, sticky="w", pady=2)
        self.entry_reinv = tk.Entry(frm_desg, font=("Arial", 10), width=12, justify="center",
                                     bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                     insertbackground="white")
        self.entry_reinv.grid(row=rw, column=1, padx=8, pady=2, sticky="w")
        rw += 1

        tk.Label(frm_desg, text="Ganancia fija:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).grid(row=rw, column=0, sticky="w", pady=2)
        self.entry_ganancia = tk.Entry(frm_desg, font=("Arial", 10), width=12, justify="center",
                                        bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                        insertbackground="white")
        self.entry_ganancia.grid(row=rw, column=1, padx=8, pady=2, sticky="w")
        rw += 1

        sep2 = tk.Frame(edit, height=1, bg=COLORS["fg_dim"])
        sep2.pack(fill="x", pady=6)

        tk.Label(edit, text="Precio de venta calculado:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack(anchor="w")
        self.lbl_pv_calculado = tk.Label(edit, text="$0",
                                          font=("Arial", 14, "bold"),
                                          bg=COLORS["bg_panel"], fg=COLORS["accent"])
        self.lbl_pv_calculado.pack(anchor="w", pady=(0, 8))

        btn_frm = tk.Frame(edit, bg=COLORS["bg_panel"])
        btn_frm.pack(fill="x", pady=4)

        tk.Button(btn_frm, text="🔄 Reajustar redondeado",
                  command=self._reajustar_redondeado,
                  bg=COLORS["blue"], fg="white",
                  font=("Arial", 9, "bold"), pady=4).pack(fill="x", pady=2)

        tk.Button(btn_frm, text="💾 Guardar cambios",
                  command=self._guardar_cambios_precio,
                  bg=COLORS["teal"], fg="white",
                  font=("Arial", 9, "bold"), pady=4).pack(fill="x", pady=2)

        self._reload_tree_precios()

    def _reload_tree_precios(self):
        for r in self.tree_precios.get_children():
            self.tree_precios.delete(r)
        for prod, val in self.cfg.precios.items():
            self.tree_precios.insert("", "end", values=(prod, f"${val:,.0f}"))

    def _on_select_precio(self, _=None):
        sel = self.tree_precios.selection()
        if not sel:
            return
        prod = self.tree_precios.item(sel)["values"][0]
        self._prod_precio_actual = prod
        self.lbl_precio_prod.config(text=prod)
        self._actualizar_desglose_precio()

    def _actualizar_desglose_precio(self):
        prod = self._prod_precio_actual
        if not prod:
            return
        vf = self.cfg.get_valores_fijos(prod)
        costo_real = self.cfg.costo_receta(prod)
        self.lbl_costo_real.config(text=f"${costo_real:,.0f}")
        self.entry_redondeado.delete(0, tk.END)
        self.entry_redondeado.insert(0, str(vf["costo_redondeado"]))
        self.entry_reinv.delete(0, tk.END)
        self.entry_reinv.insert(0, str(vf["reinversion_fija"]))
        self.entry_ganancia.delete(0, tk.END)
        self.entry_ganancia.insert(0, str(vf["ganancia_fija"]))
        self._recalcular_pv_desglose()

    def _recalcular_pv_desglose(self):
        try:
            red = float(self.entry_redondeado.get().strip() or 0)
            reinv = float(self.entry_reinv.get().strip() or 0)
            gan = float(self.entry_ganancia.get().strip() or 0)
            pv = red + reinv + gan
            self.lbl_pv_calculado.config(text=f"${pv:,.0f}")
        except ValueError:
            self.lbl_pv_calculado.config(text="Error")

    def _guardar_cambios_precio(self):
        prod = self._prod_precio_actual
        if not prod:
            messagebox.showwarning("Aviso", "Seleccioná un producto primero."); return
        try:
            red = float(self.entry_redondeado.get().strip() or 0)
            reinv = float(self.entry_reinv.get().strip() or 0)
            gan = float(self.entry_ganancia.get().strip() or 0)
        except ValueError:
            messagebox.showerror("Error", "Ingresá valores numéricos válidos."); return

        if prod not in self.cfg.valores_fijos:
            self.cfg.valores_fijos[prod] = {}
        self.cfg.valores_fijos[prod]["costo_redondeado"] = red
        self.cfg.valores_fijos[prod]["reinversion_fija"] = reinv
        self.cfg.valores_fijos[prod]["ganancia_fija"] = gan
        pv = red + reinv + gan
        self.cfg.precios[prod] = pv
        if self.cfg.save():
            self._reload_tree_precios()
            self._recalcular_pv_desglose()
            self.status_cb(f"Valores de '{prod}' guardados. PV: ${pv:,.0f}", COLORS["green"])
        else:
            messagebox.showerror("Error", "No se pudo guardar la configuración.")

    def _reajustar_redondeado(self):
        prod = self._prod_precio_actual
        if not prod:
            messagebox.showwarning("Aviso", "Seleccioná un producto primero."); return
        try:
            reajustado, nuevo, costo_real = self.cfg.reajustar_redondeado(prod)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo reajustar:\n{e}")
            return
        if reajustado:
            messagebox.showinfo("Reajuste",
                f"Costo redondeado de '{prod}' actualizado a ${nuevo:,.0f}\n"
                f"(Costo real era ${costo_real:,.0f})")
        else:
            vf = self.cfg.get_valores_fijos(prod)
            messagebox.showinfo("Sin cambios",
                f"El costo actual (${costo_real:,.0f}) no supera el\n"
                f"redondeado vigente (${vf['costo_redondeado']:,.0f}).\n"
                f"No se requiere ajuste.")
        self._actualizar_desglose_precio()

    def _build_recetas(self):
        tk.Label(self.tab_recetas,
                 text="Seleccioná un producto para ver y editar sus ingredientes y gramajes.",
                 font=("Arial", 9), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_mid"]).pack(padx=10, pady=(8, 4), anchor="w")

        pane = tk.Frame(self.tab_recetas, bg=COLORS["bg_mid"])
        pane.pack(fill="both", expand=True, padx=10, pady=4)

        izq = tk.Frame(pane, bg=COLORS["bg_mid"])
        izq.pack(side="left", fill="y")

        tk.Label(izq, text="Productos:", font=("Arial", 9, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_light"]).pack(anchor="w")
        self.lb_receta_prods = tk.Listbox(izq, font=("Arial", 9), width=30,
                                          bg=COLORS["bg_panel"], fg=COLORS["fg_light"],
                                          selectbackground=COLORS["accent"],
                                          activestyle="none", height=18)
        sb_iz = ttk.Scrollbar(izq, orient="vertical", command=self.lb_receta_prods.yview)
        self.lb_receta_prods.configure(yscrollcommand=sb_iz.set)
        self.lb_receta_prods.pack(side="left", fill="y")
        sb_iz.pack(side="left", fill="y")
        self.lb_receta_prods.bind("<<ListboxSelect>>", self._on_select_receta_prod)
        for prod in self.cfg.recetas:
            self.lb_receta_prods.insert(tk.END, prod)

        der = tk.Frame(pane, bg=COLORS["bg_mid"], padx=12)
        der.pack(side="left", fill="both", expand=True)

        tk.Label(der, text="Ingredientes de la receta:",
                 font=("Arial", 9, "bold"),
                 bg=COLORS["bg_mid"], fg=COLORS["fg_light"]).pack(anchor="w")

        cols = ("ingrediente", "cantidad", "costo")
        self.tree_receta = ttk.Treeview(der, columns=cols, show="headings", height=10)
        self.tree_receta.heading("ingrediente", text="Ingrediente")
        self.tree_receta.heading("cantidad",    text="Cantidad / Gramos")
        self.tree_receta.heading("costo",       text="Costo en receta ($)")
        self.tree_receta.column("ingrediente",  width=200, anchor="w")
        self.tree_receta.column("cantidad",     width=130, anchor="center")
        self.tree_receta.column("costo",        width=140, anchor="center")
        sb_r = ttk.Scrollbar(der, orient="vertical", command=self.tree_receta.yview)
        self.tree_receta.configure(yscrollcommand=sb_r.set)
        self.tree_receta.pack(side="left", fill="both", expand=True)
        sb_r.pack(side="left", fill="y")
        self.tree_receta.bind("<<TreeviewSelect>>", self._on_select_ingrediente)

        edit = tk.Frame(der, bg=COLORS["bg_panel"], padx=12, pady=12)
        edit.pack(fill="x", pady=(8, 0))

        tk.Label(edit, text="Ingrediente seleccionado:",
                 font=("Arial", 8, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).grid(row=0, column=0, sticky="w")
        self.lbl_ing_sel = tk.Label(edit, text="—",
                                     font=("Arial", 10, "bold"),
                                     bg=COLORS["bg_panel"], fg=COLORS["fg_light"])
        self.lbl_ing_sel.grid(row=0, column=1, sticky="w", padx=8)

        tk.Label(edit, text="Nueva cantidad:",
                 font=("Arial", 8, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_light"]).grid(row=1, column=0, sticky="w", pady=4)
        self.entry_nueva_cant = tk.Entry(edit, font=("Arial", 11), width=10,
                                          justify="center",
                                          bg=COLORS["bg_entry"], fg=COLORS["fg_light"],
                                          insertbackground="white")
        self.entry_nueva_cant.grid(row=1, column=1, padx=8, pady=4, sticky="w")

        tk.Button(edit, text="Actualizar gramaje",
                  command=self._actualizar_gramaje,
                  bg=COLORS["teal"], fg="white",
                  font=("Arial", 9, "bold")).grid(row=1, column=2, padx=8)

        self.lbl_total_receta = tk.Label(der, text="",
                                          font=("Arial", 9, "bold"),
                                          bg=COLORS["bg_mid"], fg=COLORS["fg_light"])
        self.lbl_total_receta.pack(anchor="w", pady=(4, 0))

        self._prod_receta_actual = None

    def _on_select_receta_prod(self, _=None):
        sel = self.lb_receta_prods.curselection()
        if not sel:
            return
        prod = self.lb_receta_prods.get(sel[0])
        self._prod_receta_actual = prod
        for r in self.tree_receta.get_children():
            self.tree_receta.delete(r)
        total = 0
        for ing, cant in self.cfg.recetas.get(prod, {}).items():
            costo = cant * self.cfg.costos.get(ing, 0)
            total += costo
            self.tree_receta.insert("", "end", values=(ing, f"{cant:g}", f"${costo:,.0f}"))
        vf = self.cfg.get_valores_fijos(prod)
        self.lbl_total_receta.config(
            text=f"Costo total de receta: ${total:,.0f}  →  Redondeado vigente: ${vf['costo_redondeado']:,.0f}")

    def _on_select_ingrediente(self, _=None):
        sel = self.tree_receta.selection()
        if not sel:
            return
        ing, cant, _ = self.tree_receta.item(sel)["values"]
        self.lbl_ing_sel.config(text=ing)
        self.entry_nueva_cant.delete(0, tk.END)
        self.entry_nueva_cant.insert(0, str(cant))

    def _actualizar_gramaje(self):
        if not self._prod_receta_actual:
            messagebox.showwarning("Aviso", "Seleccioná un producto primero."); return
        sel = self.tree_receta.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccioná un ingrediente."); return
        ing = self.tree_receta.item(sel)["values"][0]
        try:
            nueva = float(self.entry_nueva_cant.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Ingresá un número válido."); return
        self.cfg.recetas[self._prod_receta_actual][ing] = nueva
        self._on_select_receta_prod()
        self.status_cb(
            f"Gramaje de '{ing}' en '{self._prod_receta_actual}' -> {nueva:g} (sin guardar aún)",
            COLORS["accent"])

    def guardar_config(self):
        if self.cfg.save():
            self.refresh_cb()
            messagebox.showinfo("Guardado", "Configuración guardada correctamente.\nLos cambios ya están activos en Pedidos.")
            self.status_cb("Configuración guardada y aplicada", COLORS["green"])


# ══════════════════════════════════════════════════════════════════════════════
#  PESTAÑA 4 — ESTADÍSTICAS DE VENTAS
# ══════════════════════════════════════════════════════════════════════════════
class EstadisticasTab(tk.Frame):

    COLORES_BARRA = ["#FF8F00", "#E65100", "#C62828", "#6A1B9A", "#1565C0",
                     "#00695C", "#2E7D32", "#424242", "#37474F", "#455A64"]

    def __init__(self, parent, status_cb):
        super().__init__(parent, bg=COLORS["bg_mid"])
        self.status_cb = status_cb
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=COLORS["bg_mid"], pady=6)
        hdr.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(hdr, text="Estadísticas", font=("Segoe UI", 14, "bold"),
                 fg=COLORS["fg_light"], bg=COLORS["bg_mid"]).pack(side="left")
        tk.Button(hdr, text="🔄 Actualizar", command=self.actualizar,
                  bg=COLORS["blue"], fg="white",
                  font=("Segoe UI", 9, "bold"), padx=10).pack(side="right")

        # Card: Total acumulado
        card_total = tk.Frame(self, bg=COLORS["bg_panel"], padx=16, pady=12)
        card_total.pack(fill="x", padx=16, pady=(8, 0))
        tk.Label(card_total, text="Ventas Acumuladas",
                 font=("Segoe UI", 9, "bold"),
                 fg=COLORS["fg_dim"], bg=COLORS["bg_panel"]).pack(anchor="w")
        self.lbl_total = tk.Label(card_total, text="Cargando...",
                                  font=("Segoe UI", 24, "bold"),
                                  fg=COLORS["accent"], bg=COLORS["bg_panel"])
        self.lbl_total.pack(fill="x", anchor="w")

        # Card: Grafico
        frm = tk.Frame(self, bg=COLORS["bg_panel"], padx=2, pady=2)
        frm.pack(fill="both", expand=True, padx=16, pady=(8, 10))

        hdr_chart = tk.Frame(frm, bg=COLORS["bg_panel"])
        hdr_chart.pack(fill="x", padx=10, pady=(6, 2))
        tk.Label(hdr_chart, text="Productos más vendidos",
                 font=("Segoe UI", 10, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_panel"]).pack(side="left")

        self.canvas = tk.Canvas(frm, bg=COLORS["bg_dark"],
                                highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=(8, 0), pady=(0, 8))

        self.actualizar()

    def actualizar(self):
        self.canvas.delete("all")
        data = cargar_ventas_historicas()
        total = data.get("total", 0)
        prods = data.get("productos", {})
        self.lbl_total.config(text=f"${total:,.0f}" if total > 1000 else f"${total}")

        # Ordenar productos por cantidad vendida (desc)
        ranking = sorted(prods.items(), key=lambda x: -x[1])[:10]
        if not ranking:
            self.canvas.create_text(300, 60, text="Aún no hay ventas registradas.",
                                    fill=COLORS["fg_dim"],
                                    font=("Arial", 12), anchor="center")
            return

        max_cant = ranking[0][1]
        margen_izq = 190
        margen_der = 70
        alto_barra = 30
        espacio = 6

        self.canvas.update_idletasks()
        ancho_canvas = max(self.canvas.winfo_width(), 400)
        ancho_barra_max = ancho_canvas - margen_izq - margen_der

        inicio_y = 15
        for i, (prod, cant) in enumerate(ranking):
            y = inicio_y + i * (alto_barra + espacio)
            # Nombre del producto
            self.canvas.create_text(margen_izq - 8, y + alto_barra / 2,
                                    text=prod, anchor="e",
                                    fill=COLORS["fg_light"],
                                    font=("Arial", 8, "bold"))
            # Barra
            ancho = int((cant / max_cant) * ancho_barra_max) if max_cant > 0 else 0
            color = self.COLORES_BARRA[i % len(self.COLORES_BARRA)]
            self.canvas.create_rectangle(margen_izq, y,
                                         margen_izq + max(ancho, 4), y + alto_barra,
                                         fill=color, outline="")
            # Cantidad al final de la barra
            self.canvas.create_text(margen_izq + max(ancho, 4) + 6, y + alto_barra / 2,
                                    text=str(cant), anchor="w",
                                    fill=COLORS["accent"],
                                    font=("Arial", 9, "bold"))

        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.status_cb("Estadísticas actualizadas", COLORS["teal"])


# ══════════════════════════════════════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"La Hogaza Madre — Sistema de Gestión v{APP_VERSION}")
        self.configure(bg=COLORS["bg_dark"])
        self.resizable(True, True)

        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = min(1200, sw - 60)
        h = min(750, sh - 80)
        self.geometry(f"{w}x{h}")
        self.minsize(900, 500)

        self.cfg = ConfigManager()
        self._apply_styles()
        self._build_ui()

    def _apply_styles(self):
        st = ttk.Style()
        st.theme_use("clam")

        st.configure(".",                   font=("Segoe UI", 10))
        st.configure("TNotebook",           background=COLORS["bg_dark"],  borderwidth=0)
        st.configure("TNotebook.Tab",       background=COLORS["bg_panel"], foreground=COLORS["fg_dim"],
                     font=("Segoe UI", 10, "bold"), padding=[16, 7])
        st.map("TNotebook.Tab",
               background=[("selected", COLORS["bg_mid"])],
               foreground=[("selected", COLORS["accent"])])

        st.configure("Menu.TCombobox",      font=("Segoe UI", 11), padding=5,
                     fieldbackground=COLORS["bg_entry"], foreground=COLORS["fg_light"],
                     arrowcolor=COLORS["accent"])
        st.map("Menu.TCombobox",
               fieldbackground=[("readonly", COLORS["bg_entry"])],
               foreground=[("readonly", COLORS["fg_light"])])

        st.configure("Treeview",            background=COLORS["bg_dark"],
                     foreground=COLORS["fg_light"], rowheight=28,
                     fieldbackground=COLORS["bg_dark"], font=("Segoe UI", 9))
        st.configure("Treeview.Heading",    background=COLORS["bg_panel"],
                     foreground=COLORS["accent"], font=("Segoe UI", 9, "bold"))
        st.map("Treeview",                 background=[("selected", COLORS["bg_entry"])])

        self.option_add("*TCombobox*Listbox.font",             ("Segoe UI", 11))
        self.option_add("*TCombobox*Listbox.background",       COLORS["bg_panel"])
        self.option_add("*TCombobox*Listbox.foreground",       COLORS["fg_light"])
        self.option_add("*TCombobox*Listbox.selectBackground", COLORS["accent"])
        self.option_add("*TCombobox*Listbox.selectForeground", "white")

    def _crear_btn_nav(self, parent, icon, text, tab):
        frm = tk.Frame(parent, bg=COLORS["bg_panel"], cursor="hand2")
        frm.pack(fill="x", padx=0, pady=0)

        # Indicator bar
        ind = tk.Frame(frm, width=3, bg=COLORS["bg_panel"])
        ind.pack(side="left", fill="y")

        lbl = tk.Label(frm, text=f"  {icon}  {text}",
                       font=("Segoe UI", 11), anchor="w",
                       bg=COLORS["bg_panel"], fg=COLORS["fg_dim"],
                       padx=12, pady=10)
        lbl.pack(side="left", fill="x", expand=True)

        def on_enter(e):
            if self._tab_actual is not tab:
                frm.config(bg=COLORS["bg_entry"])
                lbl.config(bg=COLORS["bg_entry"])
                ind.config(bg=COLORS["bg_entry"])

        def on_leave(e):
            if self._tab_actual is not tab:
                frm.config(bg=COLORS["bg_panel"])
                lbl.config(bg=COLORS["bg_panel"])
                ind.config(bg=COLORS["bg_panel"])

        def on_click(e):
            self._mostrar_tab(tab)

        frm.bind("<Enter>", on_enter)
        frm.bind("<Leave>", on_leave)
        frm.bind("<Button-1>", on_click)
        lbl.bind("<Button-1>", on_click)
        ind.bind("<Button-1>", on_click)

        return {"frame": frm, "label": lbl, "indicator": ind}

    def _mostrar_tab(self, tab):
        if self._tab_actual:
            self._tab_actual.pack_forget()
        self._tab_actual = tab
        tab.pack(fill="both", expand=True)
        for t, btn in self._nav_buttons.items():
            active = t is tab
            btn["frame"].config(bg=COLORS["bg_entry"] if active else COLORS["bg_panel"])
            btn["label"].config(bg=COLORS["bg_entry"] if active else COLORS["bg_panel"],
                                fg=COLORS["accent"] if active else COLORS["fg_dim"])
            btn["indicator"].config(bg=COLORS["accent"] if active else COLORS["bg_panel"])

    def _build_ui(self):
        # ── TOP BAR ──
        hdr = tk.Frame(self, bg=COLORS["bg_mid"], pady=0)
        hdr.pack(fill="x")
        inner = tk.Frame(hdr, bg=COLORS["bg_mid"], pady=10)
        inner.pack(fill="x", padx=18)
        tk.Label(inner, text="LA HOGAZA MADRE",
                 font=("Segoe UI", 16, "bold"),
                 fg=COLORS["accent"], bg=COLORS["bg_mid"]).pack(side="left")
        tk.Label(inner, text=f"Gestión v{APP_VERSION}",
                 font=("Segoe UI", 9), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_mid"]).pack(side="left", padx=(10, 0))
        tk.Label(inner, text=datetime.now().strftime("%a %d/%m/%Y").upper(),
                 font=("Segoe UI", 8, "bold"), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_mid"]).pack(side="right", padx=(0, 4))
        tk.Frame(hdr, height=2, bg=COLORS["accent"]).pack(fill="x")

        # ── STATUS BAR (must be before tabs, they call _status on init) ──
        self.lbl_status = tk.Label(self, text="Sistema listo.",
                                    font=("Segoe UI", 9), fg=COLORS["fg_dim"],
                                    bg=COLORS["bg_dark"], anchor="w", padx=16, pady=4)
        self.lbl_status.pack(fill="x", side="bottom")

        # ── MAIN CONTENT: SIDEBAR + CONTENT ──
        main = tk.Frame(self, bg=COLORS["bg_dark"])
        main.pack(fill="both", expand=True)

        # Sidebar
        sidebar = tk.Frame(main, bg=COLORS["bg_panel"], width=150)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Logo area
        logo_frm = tk.Frame(sidebar, bg=COLORS["bg_panel"], pady=14)
        logo_frm.pack(fill="x")
        tk.Label(logo_frm, text="🍕", font=("Segoe UI", 20),
                 bg=COLORS["bg_panel"], fg=COLORS["accent"]).pack()
        tk.Label(logo_frm, text="Menú", font=("Segoe UI", 8, "bold"),
                 bg=COLORS["bg_panel"], fg=COLORS["fg_dim"]).pack()
        tk.Frame(sidebar, height=1, bg=COLORS["bg_entry"]).pack(fill="x", padx=12, pady=4)

        # Content area
        self.content = tk.Frame(main, bg=COLORS["bg_dark"])
        self.content.pack(side="left", fill="both", expand=True)

        # Create tabs
        self.tab_pedidos  = PedidosTab(self.content, self.cfg, self._status)
        self.tab_stock    = StockTab(self.content, self.cfg, self._status)
        self.tab_config   = ConfigTab(self.content, self.cfg, self._status, self._on_config_saved)
        self.tab_estadisticas = EstadisticasTab(self.content, self._status)

        # Navigation buttons
        self._tab_actual = None
        self._nav_buttons = {}
        nav_items = [
            ("📋", "Pedidos",    self.tab_pedidos),
            ("📦", "Stock",      self.tab_stock),
            ("⚙️", "Config.",    self.tab_config),
            ("📊", "Stats",      self.tab_estadisticas),
        ]
        for icon, text, tab in nav_items:
            btn = self._crear_btn_nav(sidebar, icon, text, tab)
            self._nav_buttons[tab] = btn

        # Version en sidebar
        tk.Frame(sidebar, height=1, bg=COLORS["bg_entry"]).pack(fill="x", padx=12, pady=4)
        tk.Label(sidebar, text=f"v{APP_VERSION}",
                 font=("Segoe UI", 8), fg=COLORS["fg_dim"],
                 bg=COLORS["bg_panel"]).pack(side="bottom", pady=8)

        self._mostrar_tab(self.tab_pedidos)

        self.bind("<Return>", lambda e: self.tab_pedidos.agregar_producto())
        self.bind("<F5>",     lambda e: self.tab_pedidos.guardar_pedido())

    def _status(self, msg, color=None):
        self.lbl_status.config(text=msg, fg=color or COLORS["fg_dim"])
        self.after(6000, lambda: self.lbl_status.config(
            text="Sistema listo.", fg=COLORS["fg_dim"]))

    def _on_config_saved(self):
        self.tab_pedidos.cfg = self.cfg
        self.tab_pedidos.refresh_combo_productos()
        self.tab_stock.cargar_stock()
        if self.tab_config._prod_receta_actual:
            self.tab_config._on_select_receta_prod()


# ══════════════════════════════════════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
